import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import numbers

# Definir los titulos a buscar
titulos = ["AE38", "AE38D", "AL29", "AL29D", "AL30", "AL30C", "AL30D", "AL35", "AL35D", "AL41", "AL41D",
           "BA37D", "BA37E", "BB37D", "BB37E", "BC37D", "BDC24", "BDC28", "BP21", "BT02", "BT03",
           "CEDI", "CO26", "CO26D", "CUAP", "DICP", "GD29", "GD29D", "GD30", "GD30C", "GD30D", "GD35",
           "GD35D", "GD38", "GD38D", "GD41", "GD41D", "GD46", "GD46D", "GE29", "GE30", "GE41", "GE46",
           "NDT25", "PAP0", "PARP", "PAY0", "PBA25", "PMM29", "PR13", "SA24D", "T2X3", "T2X4", "TDJ23",
           "TDL23", "TDS23", "TO23", "TO26", "TS27", "TSCH9", "TV24", "TVPA", "TVPE", "TVPP", "TVPY",
           "TVY0", "TX24", "TX25", "TX26", "TX26D", "TX28", "TX28D", "TY05", "YCA6O", "YPCUO"]

# Crear una expresión regular de los títulos
pattern = re.compile("|".join(titulos))

# Listar todos los archivos txt en el directorio actual
archivos = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.txt')]

data_pesos = {'BONOS': [], 'PESOS': []}
data_dolares = {'BONOS': [], 'DOLARES': []}

for archivo in archivos:
    with open(archivo, 'r') as f:
        lines = f.readlines()
    for i in range(len(lines)-1):
        if pattern.fullmatch(lines[i].strip()):
            # Convertir los números de cadena a formato flotante
            precio = lines[i+1].strip().replace('.', '').replace(',', '.')
            precio = float(precio) if precio != '-' else None
            if lines[i].strip().endswith('D'):
                data_dolares['BONOS'].append(lines[i].strip())
                data_dolares['DOLARES'].append(precio)
            else:
                data_pesos['BONOS'].append(lines[i].strip())
                data_pesos['PESOS'].append(precio)

# Crear un dataframe con los datos
df_pesos = pd.DataFrame(data_pesos)
df_dolares = pd.DataFrame(data_dolares)

# Ordenar la columna 'PESOS' y 'DOLARES' de menor a mayor
df_pesos.sort_values('PESOS', inplace=True)
df_dolares.sort_values('DOLARES', inplace=True)

# Generar un nombre de archivo basado en la fecha y hora actuales
fecha_hora = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

# Nombre del archivo en mayúsculas
nombre_archivo = f'ACTUALIZACION {fecha_hora}.xlsx'

# Crear un ExcelWriter para guardar el dataframe
with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
    df_pesos.to_excel(writer, sheet_name='COTIZACIONES PESOS', index=False)
    df_dolares.to_excel(writer, sheet_name='COTIZACIONES DOLARES', index=False)

# Cargar el archivo creado con openpyxl
book = load_workbook(nombre_archivo)

# Asignar el formato de moneda a las columnas PESOS y DOLARES
for sheet in book.sheetnames:
    for column in ['B', 'C']:
        for cell in book[sheet][column]:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

# Guardar los cambios
book.save(nombre_archivo)
